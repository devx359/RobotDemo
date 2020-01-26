from openpyxl import load_workbook
#class Excel :
def  load(self, filepath):
     global sheet
     global wb
     wb = load_workbook(filepath);
     print(wb.sheetnames)
     sheet = wb["Sheet1"]
     print("sheet is of type "+str(type(sheet)))

def getCellValue(self,cellName):
    return sheet[cellName].value

def getRangeOfValues(self,cellRange):
    tup = sheet[cellRange]
    return tup

def getmultipleValues(self):
    for value in sheet.iter_rows(min_row=1,max_row=2,min_col=1,max_col=2,values_only=True):
        print(value)






