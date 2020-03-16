from openpyxl import Workbook, load_workbook


class  excelUtil:

    def  __init__(self):
        print("inside constructor")

    def  openExcel(self, excelFilePath ):
        global wb
        global ws
        try:
            wb = load_workbook(excelFilePath) #opens the excel workbook file
            print("Excel file opened")

        except  Exception as e:
            print("Failed to open excel file "+ str(e));
            raise  e


    def  getCellValueByCellNumber(self ,sheetName,cellNumber):
        try:
            ws = wb[sheetName]  # open the desired sheet
            c = ws[cellNumber]
            #print(type(c))
            print(c.value)
            return  (c.value)
        except  Exception as e:
            print("Error when fetching cell value "+ str(e))
            return None

    def  getCellsValueByCellRange(self, sheetName , cellRange):
        ll = []
        ws = wb[sheetName]
        crange = ws[cellRange] #gets tuple of cells
        #print(type(crange))

        for var in crange:  #tuples contain tuple of size 1 containing a cell object
            print (var[0].value)
            ll.append(var[0].value)  #var[0] = cell object contained in tuple 0 index

        return ll

    def  readColoumn(self, sheetName , coloumnName):
        ws = wb[sheetName]
        coloumn = ws[coloumnName]
        #print(type(coloumn))
        #print(len(coloumn))
        ll = []
        for cell in coloumn:
            print(cell.value)
            ll.append(cell.value)

        return (ll)


    def  closeExcel(self):
        wb.close()
        print("closed the excel")


